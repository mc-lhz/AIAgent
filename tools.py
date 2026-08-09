from openpyxl import load_workbook, Workbook
import os
import subprocess
def stop(ai_result: str):
    """
    停止执行，展示结果给用户
    执行完所有任务完成后请输出stop('需要展示给用户的结果')停止执行
    参数:
        ai_result: 要展示给用户的结果
    """
    global continueFlag
    print(ai_result)
    return '<stop>'
def runBash(command: str) -> str:
    """
    执行Bash命令并返回输出
    
    参数:
        command: 要执行的Bash命令，如 'ls'
    
    返回:
        str: 命令执行结果
    """
    result = subprocess.run(command, shell=True, capture_output=True, text=True)
    if result.stderr:
        return f"错误：{result.stderr.strip()}"
    
    # 如果 stdout 有内容，返回 stdout
    if result.stdout:
        return result.stdout.strip()
    
    # 两者都空，说明命令无输出
    return "（命令执行成功，但无输出）"
def runPowerShell(command: str) -> str:
    """
    执行PowerShell命令并返回输出
    
    参数:
        command: 要执行的PowerShell命令，如 'Get-Process'
    
    返回:
        str: 命令执行结果
    """
    result = subprocess.run(f"powershell -Command \"{command}\"", capture_output=True, text=True)
    if result.stderr:
        return f"错误：{result.stderr.strip()}"
    
    # 如果 stdout 有内容，返回 stdout
    if result.stdout:
        return result.stdout.strip()
    
    # 两者都空，说明命令无输出
    return "（命令执行成功，但无输出）"
    
def excelOperation(filename: str, operations: list) -> str:
    """
    对Excel文件执行批量单元格操作
    
    参数:
        filename: 文件名（含.xlsx后缀），如 '成绩表.xlsx'
        operations: 操作列表，每个操作格式为：
            [
                {"method": "read", "x": "A", "y": 2},              # 读取A2
                {"method": "write", "x": "B", "y": 3, "value": "Hello"}  # 写入B3
            ]
    
    返回:
        str: 操作结果汇总
    """
    results = []
    # 如果文件不存在，创建一个新的
    if not os.path.exists(f"{filename}"):
        # 创建新的工作簿
        wb = Workbook()
        results.append({
            "result": f"文件 {filename} 不存在，已创建一个新的文件"
        })
    else:
        # 加载工作簿
        wb = load_workbook(f"{filename}")
    ws = wb.active
    
    
    
    for op in operations:
        method = op.get("method")
        col = op.get("x")      # 列字母，如 'A', 'B'
        row = op.get("y")      # 行号，从1开始
        value = op.get("value")
        
        if method == "read":
            cellValue = ws[f"{col}{row}"].value
            results.append({
                "operation": op,
                "result": f"{col}{row} = {cellValue}" if cellValue is not None else f"{col}{row} = (空)"
            })
        
        elif method == "write":
            ws[f"{col}{row}"] = value
            results.append({
                "operation": op,
                "result": f"已写入 {col}{row} = {value}"
            })
        
        else:
            results.append({
                "operation": op,
                "result": f"不支持的method: {method}"
            })
    
        # 保存文件
    wb.save(f"{filename}")
    
    # 返回汇总信息
    summary = "\n".join([r["result"] for r in results])
    return f"操作完成！\n{summary}"